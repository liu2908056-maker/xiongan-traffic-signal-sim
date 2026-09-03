"""Build the 20-junction Xiong'an CityFlow scenario from supplied workbooks.

The geometry is a schematic 4 x 5 representation of the Rongdong road network.
It preserves the direction and observed turning volumes of every supplied sheet;
it is not a substitute for a surveyed OSM/netedit lane-level network.
"""
from __future__ import annotations

import argparse
import json
import random
import re
from collections import Counter
from datetime import time
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "xiong_an_20"
APPROACHES = ("W", "E", "N", "S", "NE", "NW", "SE", "SW")
CARDINAL = ("W", "E", "N", "S")
MOVEMENTS = ("left", "through", "right")
# Diagonal labels must be recognized before cardinal labels; otherwise a
# "northeast approach" is silently misread as either north or east.
CHINESE = (("东北", "NE"), ("西北", "NW"), ("东南", "SE"), ("西南", "SW"), ("西", "W"), ("东", "E"), ("北", "N"), ("南", "S"))
CONTROL_GROUP = {"W": "W", "E": "E", "N": "N", "S": "S", "NW": "N", "NE": "N", "SW": "S", "SE": "S"}
ANGLE = ("N", "NE", "E", "SE", "S", "SW", "W", "NW")
DEST = {approach: {"left": ANGLE[(ANGLE.index(approach) + 2) % 8], "through": ANGLE[(ANGLE.index(approach) + 4) % 8], "right": ANGLE[(ANGLE.index(approach) + 6) % 8]} for approach in APPROACHES}
OPPOSITE = {approach: ANGLE[(ANGLE.index(approach) + 4) % 8] for approach in APPROACHES}

# Approximate relative placement from supplied AMap screenshots/landmarks. North is +Y.
GRID = ((7, 1, 4, 8, 10), (5, 6, 3, 2, 9), (13, 11, 12, 16, 20), (14, 17, 15, 18, 19))
POSITION = {node: (col, 3 - row) for row, values in enumerate(GRID) for col, node in enumerate(values)}
# Landmark names are taken from the supplied AMap/high-resolution filenames and
# the existing intersection-2 location note. Unknown locations remain explicit
# approximate labels instead of being presented as surveyed OSM coordinates.
LABELS = {
    1: "雄州路711号附近", 2: "津海大街-奥威路（容和塔北侧）", 3: "铃铛阁大街2号附近", 4: "将台路217号附近", 5: "雄州镇城区西侧",
    6: "雄州镇城区中心", 7: "中心大街16号附近", 8: "雄州路788号附近", 9: "铃铛阁大街221号附近", 10: "保静公路G336附近",
    11: "容东片区西北侧", 12: "容东片区北部", 13: "小康路2号附近", 14: "容东片区西南侧", 15: "容东片区中心南侧",
    16: "容东片区东北侧", 17: "津保路1号附近（西侧）", 18: "小康路6号附近", 19: "津保路1号附近（东侧）", 20: "小康路2号附近",
}
SHORT_LABELS = {
    1: "雄州路711", 2: "津海-奥威", 3: "铃铛阁2", 4: "将台路217", 5: "雄州城区西", 6: "雄州城区中", 7: "中心大街16", 8: "雄州路788", 9: "铃铛阁221", 10: "保静G336",
    11: "容东西北", 12: "容东北部", 13: "小康路2", 14: "容东西南", 15: "容东中心南", 16: "容东北侧", 17: "津保路西", 18: "小康路6", 19: "津保路东", 20: "小康路2",
}
# Slightly irregular spacing reflects the road directions visible in the
# supplied imagery while preserving the requested regular 4x5 topology.
COORDS = {
    7: (0.0, 1420.0), 1: (410.0, 1450.0), 4: (830.0, 1400.0), 8: (1260.0, 1440.0), 10: (1700.0, 1410.0),
    5: (20.0, 990.0), 6: (440.0, 1020.0), 3: (850.0, 1000.0), 2: (1290.0, 1040.0), 9: (1730.0, 990.0),
    13: (-20.0, 570.0), 11: (420.0, 600.0), 12: (870.0, 560.0), 16: (1300.0, 600.0), 20: (1720.0, 560.0),
    14: (0.0, 100.0), 17: (440.0, 130.0), 15: (850.0, 100.0), 18: (1280.0, 140.0), 19: (1730.0, 100.0),
}
# A diagonal Excel approach is geometrically tied to the nearest controlled
# junction, then bends into its recorded approach side.  The virtual source is
# retained for independent demand accounting; the rendered road is continuous.
# Diagonal approaches are short, local arms.  They must not be stretched across
# the lower network: J09 remains a four-arm junction with a NE branch, while
# J18/J19 retain their four diagonal arms from the supplied workbooks.
DIAGONAL_ANCHOR: dict[int, dict[str, int]] = {}
HIDDEN_CORRIDOR_PAIRS: tuple[tuple[int, int], ...] = ()
# These empirical cardinal entry counts are visually placed on the only
# connected lower corridors instead of drawing dead-end display stubs.
FEEDER_VISUAL_LINKS = {
    "feed_15_E": "link_15_NE_18",
    "feed_16_S": "link_16_SW_18",
    "feed_20_S": "link_20_SE_19",
}


def write(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def pt(x: float, y: float) -> dict[str, float]:
    return {"x": round(x, 2), "y": round(y, 2)}


def road(road_id: str, start: str, end: str, a: tuple[float, float], b: tuple[float, float], waypoints: list[tuple[float, float]] | None = None, straight: bool = False) -> dict[str, Any]:
    dx, dy = b[0] - a[0], b[1] - a[1]; length = max(1.0, (dx * dx + dy * dy) ** 0.5)
    # Use a deterministic, small lateral offset so CityFlow retains a curved
    # polyline without making the schematic grid visually irregular.
    sign = 1 if sum(ord(c) for c in road_id) % 2 else -1; offset = min(22.0, length * 0.06) * sign
    middle = (a[0] + dx * 0.5 - dy / length * offset, a[1] + dy * 0.5 + dx / length * offset)
    points = [pt(*a), pt(*b)] if straight else ([pt(*a), *[pt(*point) for point in waypoints], pt(*b)] if waypoints else [pt(*a), pt(*middle), pt(*b)])
    return {"id": road_id, "points": points, "lanes": [
        {"width": 3.25, "maxSpeed": 13.89}, {"width": 3.25, "maxSpeed": 13.89}],
        "startIntersection": start, "endIntersection": end}


def direction_point(x: float, y: float, direction: str, distance: float) -> tuple[float, float]:
    dx, dy = {"W": (-1, 0), "E": (1, 0), "N": (0, 1), "S": (0, -1), "NE": (0.707, 0.707), "NW": (-0.707, 0.707), "SE": (0.707, -0.707), "SW": (-0.707, -0.707)}[direction]
    return x + dx * distance, y + dy * distance


def diagonal_waypoints(node: int, direction: str, x: float, y: float) -> list[tuple[float, float]]:
    """A single gentle bend makes an anchored diagonal enter from its label direction."""
    return [direction_point(x, y, direction, 112.0)]


def neighbors() -> dict[int, dict[str, int]]:
    result: dict[int, dict[str, int]] = {i: {} for i in POSITION}
    by_xy = {xy: node for node, xy in POSITION.items()}
    for node, (x, y) in POSITION.items():
        for d, delta in {"W": (-1, 0), "E": (1, 0), "N": (0, 1), "S": (0, -1)}.items():
            if (x + delta[0], y + delta[1]) in by_xy:
                result[node][d] = by_xy[x + delta[0], y + delta[1]]
    return result


def build_roadnet(observed: dict[int, set[str]]) -> tuple[dict[str, Any], dict[int, dict[str, list[str]]], dict[int, dict[str, str]], dict[int, dict[str, int]], dict[tuple[int, str], str], list[list[str]]]:
    """Create only spreadsheet-observed feeder arms and fold diagonal arms into main corridors."""
    grid_neighbors = neighbors()
    # J18/J19 use diagonal approaches only. Remove their ordinary cardinal
    # links so no east-west road is drawn in parallel with a diagonal branch.
    nbs = {node: {direction: target for direction, target in directions.items() if node not in (18, 19) and target not in (18, 19)} for node, directions in grid_neighbors.items()}
    roads: list[dict[str, Any]] = []
    intersections: list[dict[str, Any]] = []
    incoming = {node: {d: [] for d in APPROACHES} for node in POSITION}; arrivals: dict[tuple[int, str], str] = {}
    outgoing = {node: {} for node in POSITION}
    xy = COORDS.copy()
    virtual_roads: dict[str, list[str]] = {}
    corridor_roads: list[list[str]] = []

    # Local feeders preserve every observed approach count. A vehicle can then
    # cross one neighbor when the observed turn has an adjacent destination.
    for node, (x, y) in xy.items():
        for d in sorted(observed[node], key=ANGLE.index):
            virtual = f"v_{node:02d}_{d}_in"; rid = f"feed_{node:02d}_{d}"
            anchor = DIAGONAL_ANCHOR.get(node, {}).get(d)
            p = xy[anchor] if anchor else direction_point(x, y, d, 160.0)
            # The supplied workbooks explicitly name diagonal approaches at
            # J09/J18/J19.  Draw them as direct short diagonal arms; do not
            # invent a folded connector that is absent from the data table.
            fold = {"NE": "N", "NW": "W", "SE": "E", "SW": "S"}.get(d)
            if anchor:
                # The final segment enters from the exact Excel direction.
                waypoints = diagonal_waypoints(node, d, x, y)
            elif d in ("NE", "NW", "SE", "SW"):
                waypoints = None
            else:
                waypoints = [direction_point(x, y, fold, 92.0)] if fold else None
            intersections.append({"id": virtual, "point": pt(*p), "width": 0, "roads": [rid], "roadLinks": [],
                                  "trafficLight": {"roadLinkIndices": [], "lightphases": [{"time": 1, "availableRoadLinks": []}]}, "virtual": True})
            roads.append(road(rid, virtual, f"j_{node:02d}", p, (x, y), waypoints, straight=d in ("NE", "NW", "SE", "SW") and not anchor)); incoming[node][d].append(rid)
            virtual_roads[virtual] = [rid]

    # The 20 controlled intersections remain connected by their cardinal main
    # corridors; unobserved directions simply have no local external demand.
    for node, dmap in nbs.items():
        x, y = xy[node]
        for d, target in dmap.items():
            rid = f"link_{node:02d}_{d}_{target:02d}"; tx, ty = xy[target]
            roads.append(road(rid, f"j_{node:02d}", f"j_{target:02d}", (x, y), (tx, ty)))
            incoming[target][OPPOSITE[d]].append(rid)
            arrivals[node, d] = OPPOSITE[d]

    # The only in-network diagonal connections for the lower pair. J18 links
    # to J15 and J16; J19 links to J20. No other J18/J19 connection exists.
    diagonal_links = (
        (15, 18, "NE", "NW", [(1000.0, 150.0), (1135.0, 255.0)]),
        (16, 18, "SW", "NE", [(1325.0, 390.0), (1360.0, 235.0)]),
        (20, 19, "SE", "NW", [(1690.0, 360.0), (1650.0, 180.0)]),
    )
    for source, target, departure, arrival, points in diagonal_links:
        sx, sy = xy[source]; tx, ty = xy[target]
        rid = f"link_{source:02d}_{departure}_{target:02d}"
        roads.append(road(rid, f"j_{source:02d}", f"j_{target:02d}", (sx, sy), (tx, ty), points))
        incoming[target][arrival].append(rid); nbs[source][departure] = target; arrivals[source, departure] = arrival
        reverse_departure, reverse_arrival = arrival, departure
        reverse_id = f"link_{target:02d}_{reverse_departure}_{source:02d}"
        roads.append(road(reverse_id, f"j_{target:02d}", f"j_{source:02d}", (tx, ty), (sx, sy), list(reversed(points))))
        incoming[source][reverse_arrival].append(reverse_id); nbs[target][reverse_departure] = source; arrivals[target, reverse_departure] = reverse_arrival

    # Egress roads are deliberately retained for all possible movement results:
    # they keep every observed left/through/right route valid without adding a
    # visible road arm to the live monitor.
    for node, (x, y) in xy.items():
        for d in APPROACHES:
            virtual = f"v_{node:02d}_{d}_out"; rid = f"exit_{node:02d}_{d}"
            p = direction_point(x, y, d, 160.0)
            intersections.append({"id": virtual, "point": pt(*p), "width": 0, "roads": [rid], "roadLinks": [],
                                  "trafficLight": {"roadLinkIndices": [], "lightphases": [{"time": 1, "availableRoadLinks": []}]}, "virtual": True})
            roads.append(road(rid, f"j_{node:02d}", virtual, (x, y), p)); virtual_roads[virtual] = [rid]
            outgoing[node][d] = f"link_{node:02d}_{d}_{nbs[node][d]:02d}" if d in nbs[node] else rid

    phase_pairs = (("W", "through"), ("E", "through")), (("N", "through"), ("S", "through")), (("W", "left"), ("E", "left")), (("N", "left"), ("S", "left")), (("W", "left"), ("W", "through")), (("E", "left"), ("E", "through")), (("S", "left"), ("S", "through")), (("N", "left"), ("N", "through"))
    for node, (x, y) in xy.items():
        links: list[dict[str, Any]] = []; by_move: dict[tuple[str, str], list[int]] = {}
        for approach in APPROACHES:
            for movement in MOVEMENTS:
                target = DEST[approach][movement]; indices = []
                for start in incoming[node][approach]:
                    indices.append(len(links))
                    lane = 0 if movement == "left" else 1
                    links.append({"type": {"left": "turn_left", "through": "go_straight", "right": "turn_right"}[movement],
                                  "startRoad": start, "endRoad": outgoing[node][target], "direction": 0,
                                  "laneLinks": [{"startLaneIndex": lane, "endLaneIndex": lane, "points": [pt(x, y), pt(x, y)]}]})
                by_move[approach, movement] = indices
        phases = [{"time": 3, "availableRoadLinks": []}]
        for pairs in phase_pairs:
            ids: list[int] = []
            for control_approach, movement in pairs:
                for approach in APPROACHES:
                    if CONTROL_GROUP[approach] != control_approach: continue
                    ids += by_move[approach, movement]
                    if movement == "through": ids += by_move[approach, "right"]
            phases.append({"time": 30, "availableRoadLinks": ids})
        incident = [r["id"] for r in roads if r["startIntersection"] == f"j_{node:02d}" or r["endIntersection"] == f"j_{node:02d}"]
        intersections.append({"id": f"j_{node:02d}", "name": LABELS[node], "point": pt(x, y), "width": 18, "roads": incident, "roadLinks": links,
                              "trafficLight": {"roadLinkIndices": list(range(len(links))), "lightphases": phases}, "virtual": False})
    # Pair opposite directions into one visible physical corridor.
    seen: set[str] = set()
    for item in roads:
        rid = item["id"]
        if not rid.startswith("link_") or rid in seen:
            continue
        parts = rid.split("_"); source, direction, target = int(parts[1]), parts[2], int(parts[3])
        reverse = f"link_{target:02d}_{OPPOSITE[direction]}_{source:02d}"
        group = [rid] + ([reverse] if any(r["id"] == reverse for r in roads) else [])
        corridor_roads.append(group); seen.update(group)
    return {"intersections": intersections, "roads": roads}, incoming, outgoing, nbs, arrivals, corridor_roads


def seconds(value: Any) -> int:
    if isinstance(value, time): return value.hour * 3600 + value.minute * 60 + value.second
    if isinstance(value, str):
        h, m = value.strip().split(":")[:2]; return int(h) * 3600 + int(m) * 60
    raise ValueError(f"Unsupported time: {value!r}")


def parse_workbook(path: Path) -> tuple[dict[str, list[tuple[int, int, str, str, int]]], list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(path, data_only=True); sheet = wb["流量数据"]
    periods: dict[str, list[tuple[int, int, str, str, int]]] = {}; names = ("morning", "midday", "evening"); pi = 0
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row, 1).value != "统计开始时间": continue
        columns: dict[int, str] = {}; current = None
        movement_headers = [sheet.cell(row + 1, c).value for c in range(1, sheet.max_column + 1)]
        for col in range(3, sheet.max_column + 1):
            value = sheet.cell(row, col).value
            if isinstance(value, str): current = next((code for cn, code in CHINESE if f"{cn}进口" in value), current)
            if current: columns[col] = current
        demand: list[tuple[int, int, str, str, int]] = []; r = row + 2
        while r <= sheet.max_row and isinstance(sheet.cell(r, 1).value, time):
            begin, end = seconds(sheet.cell(r, 1).value), seconds(sheet.cell(r, 2).value)
            for col, approach in columns.items():
                header, value = movement_headers[col - 1], sheet.cell(r, col).value
                if not isinstance(header, str) or value in (None, ""): continue
                movement = next((m for m, cn in (("left", "左转"), ("through", "直行"), ("right", "右转")) if cn in header), None)
                if movement and float(value) > 0: demand.append((begin, end, approach, movement, int(round(float(value)))))
            r += 1
        if demand:
            zero = min(item[0] for item in demand); periods[names[pi]] = [(a-zero, b-zero, c, d, e) for a, b, c, d, e in demand]; pi += 1
    if set(periods) != set(names): raise ValueError(f"Expected three traffic periods in {path}; found {list(periods)}")
    signal_sheet = next((s for s in wb.worksheets if "配时" in s.title), None); plans: list[dict[str, Any]] = []
    if signal_sheet:
        for values in signal_sheet.iter_rows(min_row=3, values_only=True):
            if values[2] is not None:
                plans.append({"period": values[0], "range": values[1], "phase": values[2], "name": values[3], "green_s": values[4], "yellow_s": values[5], "all_red_s": values[6], "phase_total_s": values[7], "cycle_s": values[8]})
    return periods, plans


def make_flow(node: int, demand: list[tuple[int, int, str, str, int]], outgoing: dict[int, dict[str, str]], nbs: dict[int, dict[str, int]], arrivals: dict[tuple[int, str], str], seed: int) -> list[dict[str, Any]]:
    rng = random.Random(seed); result: list[dict[str, Any]] = []
    vehicle_types = (("car", 0.86, 5.0, 13.89), ("bus", 0.08, 12.0, 11.11), ("freight", 0.06, 8.5, 11.11))
    for begin, end, approach, movement, count in demand:
        for _ in range(count):
            roll = rng.random(); total = 0.0
            for name, share, length, speed in vehicle_types:
                total += share
                if roll <= total: break
            destination = DEST[approach][movement]
            # After the observed movement at its source node, continue straight
            # through connected nodes until the schematic network boundary.
            # Every adjacent pair is therefore backed by a CityFlow roadLink.
            route = [f"feed_{node:02d}_{approach}"]; current = node
            while True:
                route.append(outgoing[current][destination])
                if destination not in nbs[current]: break
                arrival = arrivals[current, destination]
                current = nbs[current][destination]
                # At an ordinary cardinal connector this leaves destination
                # unchanged. At a folded diagonal connector it follows the
                # recorded arrival arm through the receiving junction.
                destination = DEST[arrival]["through"]
            result.append({"vehicle": {"type": name, "length": length, "width": 2.0, "maxPosAcc": 2.6, "maxNegAcc": 4.5, "usualPosAcc": 2.0, "usualNegAcc": 4.5, "minGap": 2.5, "maxSpeed": speed, "headwayTime": 1.2}, "route": route, "interval": 1.0, "startTime": rng.randrange(begin, end), "endTime": begin})
    result.sort(key=lambda item: item["startTime"]); return result


def main() -> None:
    ap = argparse.ArgumentParser(); ap.add_argument("--source-dir", type=Path, required=True); ap.add_argument("--output-dir", type=Path, default=OUT); args = ap.parse_args()
    files = sorted(args.source_dir.glob("*/路口数据/demo_*流量和交叉口配时方案.xlsx"), key=lambda p: int(re.search(r"demo_(\d+)", p.name).group(1)))
    if len(files) != 20: raise ValueError(f"Expected 20 workbooks, found {len(files)} in {args.source_dir}")
    parsed: dict[int, tuple[Path, dict[str, list[tuple[int, int, str, str, int]]], list[dict[str, Any]]]] = {}
    observed_approaches: dict[int, set[str]] = {}
    for workbook in files:
        node = int(re.search(r"demo_(\d+)", workbook.name).group(1)); periods, plans = parse_workbook(workbook)
        parsed[node] = (workbook, periods, plans)
        observed_approaches[node] = {approach for demand in periods.values() for _, _, approach, _, _ in demand}
    roadnet, incoming, outgoing, nbs, arrivals, corridor_roads = build_roadnet(observed_approaches); write(args.output_dir / "roadnet.json", roadnet)
    combined = {name: [] for name in ("morning", "midday", "evening")}; manifest = []; signal_plans = {}; audit: dict[str, Any] = {}
    for node, (workbook, periods, plans) in parsed.items():
        signal_plans[str(node)] = plans; detail = {}; audit[str(node)] = {"source": str(workbook), "periods": {}}
        for index, (period, demand) in enumerate(periods.items()):
            flow = make_flow(node, demand, outgoing, nbs, arrivals, 20260816 + node * 10 + index); combined[period] += flow
            by_approach = Counter(); by_turn = Counter()
            for _, _, approach, movement, count in demand:
                by_approach[approach] += count; by_turn[f"{approach}_{movement}"] += count
            detail[period] = {"vehicles": len(flow), "approach_vehicles": dict(by_approach)}
            audit[str(node)]["periods"][period] = {"excel_total": sum(by_approach.values()), "generated_total": len(flow), "approach_vehicles": dict(by_approach), "turning_vehicles": dict(by_turn)}
        manifest.append({"intersection": node, "grid_position": POSITION[node], "source": str(workbook), "periods": detail})
    for name, flow in combined.items(): flow.sort(key=lambda item: item["startTime"]); write(args.output_dir / f"flow_{name}.json", flow)
    write(args.output_dir / "manifest.json", manifest); write(args.output_dir / "source_signal_plans.json", signal_plans); write(args.output_dir / "excel_audit.json", audit)
    write(args.output_dir / "cityflow.config.json", {"interval": 1.0, "seed": 20260816, "dir": ".", "roadnetFile": "roadnet.json", "flowFile": "flow_morning.json", "rlTrafficLight": True, "saveReplay": False, "roadnetLogFile": "logs/roadnet.log", "replayLogFile": "logs/replay.txt"})
    write(args.output_dir / "topology.json", {"grid": GRID, "positions": POSITION, "coordinates_m": COORDS, "labels": LABELS, "short_labels": SHORT_LABELS, "observed_approaches": {str(node): sorted(values, key=ANGLE.index) for node, values in observed_approaches.items()}, "neighbors": nbs, "corridor_roads": corridor_roads, "hidden_corridor_pairs": HIDDEN_CORRIDOR_PAIRS, "feeder_visual_links": FEEDER_VISUAL_LINKS, "modeling_note": "Excel approach directions and turning volumes are authoritative. Curved, folded connectors are schematic links between the 20 controlled nodes, not a surveyed OSM/netedit export."})
    print(json.dumps({"intersections": 20, "roads": len(roadnet["roads"]), "flows": {p: len(v) for p, v in combined.items()}}, ensure_ascii=False))


if __name__ == "__main__": main()
